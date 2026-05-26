#!/usr/bin/env python3
"""
Insentek OpenAPI Excel Export Script

依赖: pip install openpyxl

Usage:
    python export_excel.py --token TOKEN --sn SN --range START,END --output file.xlsx [--include-params moisture,temperature]
"""

import argparse
import json
import os
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl。请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

API_BASE_URL = os.environ.get("INSENTEK_API_BASE", "https://openapi.ecois.info")
MAX_RANGE_DAYS = 365
MAX_HISTORY_YEARS = 3
MAX_EXPORT_ROWS = 50000


def api_request(path, headers=None, params=None):
    url = f"{API_BASE_URL}{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url)
    if headers:
        for k, v in headers.items():
            req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8")
        try:
            err = json.loads(body)
        except json.JSONDecodeError:
            err = {"message": body}
        return {"_http_error": True, "status": e.code, "error": err}
    except Exception as e:
        return {"_http_error": True, "status": 0, "error": {"message": str(e)}}


def validate_range(range_str):
    parts = range_str.split(",")
    if len(parts) != 2:
        return False, "格式错误，应为 YYYYMMDD,YYYYMMDD", None, None
    try:
        start = datetime.strptime(parts[0], "%Y%m%d").date()
        end = datetime.strptime(parts[1], "%Y%m%d").date()
    except ValueError:
        return False, "日期格式错误", None, None

    today = date.today()
    if start > end:
        return False, "开始日期不能晚于结束日期", start, end

    span_days = (end - start).days + 1
    if span_days > MAX_RANGE_DAYS:
        return False, f"跨度 {span_days} 天超过限制 {MAX_RANGE_DAYS} 天", start, end

    max_history = today - timedelta(days=365 * MAX_HISTORY_YEARS)
    if end < max_history:
        return False, f"超过最大历史回溯 {MAX_HISTORY_YEARS} 年", start, end

    return True, "", start, end


def query_data(token, sn, range_str, include_params=None):
    ok, err_msg, _, _ = validate_range(range_str)
    if not ok:
        return {"_validation_error": True, "message": err_msg}

    params = {"range": range_str}
    if include_params:
        params["includeParameters"] = include_params

    return api_request(f"/v3/device/{sn}/data", headers={"Authorization": token}, params=params)


def flatten_data(api_response):
    rows = []
    for item in api_response.get("list", []):
        ts = item.get("timestamp")
        dt = item.get("datetime")
        values = item.get("values", {})
        for node, params in values.items():
            for param, value in params.items():
                rows.append({
                    "timestamp": ts,
                    "datetime": dt,
                    "node": node,
                    "parameter": param,
                    "value": value,
                })
    return rows


DRY_RUN_PREVIEW_ROWS = 5


def dry_run_summary(result, sn, range_str):
    """生成 --dry-run 预览输出。"""
    if result.get("_validation_error") or result.get("_http_error"):
        print(json.dumps({"success": False, "error": result.get("message") or result.get("error")}, ensure_ascii=False, indent=2))
        sys.exit(1)

    total = result.get("total", 0)
    flat = flatten_data(result)
    nodes = sorted(set(r["node"] for r in flat))
    params = sorted(set(r["parameter"] for r in flat))
    preview = flat[:DRY_RUN_PREVIEW_ROWS]

    output = {
        "dry_run": True,
        "total": total,
        "time_range": {
            "start": range_str.split(",")[0] if "," in range_str else "",
            "end": range_str.split(",")[1] if "," in range_str else ""
        },
        "fields": {
            "nodes": nodes,
            "parameters": params
        },
        "preview": preview,
        "message": f"Dry run: {total} records, {len(nodes)} nodes, {len(params)} parameters. First {len(preview)} rows shown. Use without --dry-run to export full data."
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))


def export_excel(token, sn, range_str, output_path, include_params=None, dry_run=False):
    result = query_data(token, sn, range_str, include_params)
    if result.get("_validation_error") or result.get("_http_error"):
        print(json.dumps({"success": False, "error": result.get("message") or result.get("error")}, ensure_ascii=False, indent=2))
        sys.exit(1)

    total = result.get("total", 0)
    if total > MAX_EXPORT_ROWS:
        print(json.dumps({
            "success": False,
            "error": f"数据量 {total} 条超过导出上限 {MAX_EXPORT_ROWS} 条，建议缩小时间范围或分批导出。"
        }, ensure_ascii=False, indent=2))
        sys.exit(1)

    if dry_run:
        dry_run_summary(result, sn, range_str)
        return

    rows = flatten_data(result)
    if not rows:
        print(json.dumps({"success": True, "total": 0, "file": output_path, "message": "无数据"}, ensure_ascii=False, indent=2))
        return

    wb = Workbook()

    # Sheet 1: 原始数据
    ws1 = wb.active
    ws1.title = "原始数据"
    headers = ["时间戳", "时间", "节点", "参数", "数值"]
    ws1.append(headers)
    for r in rows:
        ws1.append([r["timestamp"], r["datetime"], r["node"], r["parameter"], r["value"]])

    # 样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = thin_border

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws1[col_letter]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws1.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Sheet 2: 统计摘要
    ws2 = wb.create_sheet("统计摘要")
    ws2.append(["指标", "样本数", "平均值", "最小值", "最大值"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    stats = {}
    for r in rows:
        key = f"{r['node']} - {r['parameter']}"
        if key not in stats:
            stats[key] = []
        try:
            stats[key].append(float(r["value"]))
        except (TypeError, ValueError):
            pass

    for key, vals in stats.items():
        if vals:
            ws2.append([
                key,
                len(vals),
                round(sum(vals) / len(vals), 3),
                round(min(vals), 3),
                round(max(vals), 3),
            ])

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = thin_border

    for col_idx in range(1, 6):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws2[col_letter]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws2.column_dimensions[col_letter].width = min(max_length + 2, 40)

    wb.save(output_path)

    print(json.dumps({
        "success": True,
        "total": total,
        "file": str(Path(output_path).resolve()),
        "message": f"成功导出 {total} 条数据到 {output_path}"
    }, ensure_ascii=False, indent=2))


def main():
    global API_BASE_URL
    parser = argparse.ArgumentParser(description="Insentek Excel Export")
    parser.add_argument("--token", required=True)
    parser.add_argument("--sn", required=True)
    parser.add_argument("--range", required=True, help="格式: YYYYMMDD,YYYYMMDD")
    parser.add_argument("--output", required=True)
    parser.add_argument("--include-params", default=None)
    parser.add_argument("--api-base", default=API_BASE_URL)
    parser.add_argument("--dry-run", action="store_true", help="仅预览数据摘要，不写入文件")
    args = parser.parse_args()

    API_BASE_URL = args.api_base

    export_excel(args.token, args.sn, args.range, args.output, args.include_params, args.dry_run)


if __name__ == "__main__":
    main()
